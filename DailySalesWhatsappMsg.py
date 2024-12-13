# Generate Text file from master invertory.xlsx to text file
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta
# Load the workbook and specify the sheet name
file_path = "C:/All Sales/Sales/Master Inventory.xlsx"  # Path to your Excel file
workbook = load_workbook(file_path)
sheet = workbook.active





# Get yesterday's date
yesterday_date = datetime.now() - timedelta(days=1)
formatted_yesterday_date = yesterday_date.strftime("%d %b %y %A")

# Step 2: Initialize result_list as an empty list
result_list = [{"code":" ","name":formatted_yesterday_date,"value":" "},{"code":" ","name":"AMAZON SALES SUMMARY","value":" "},{"code":" ","name":" ","value":" "}]
totalProduct=0      #total product value for removing code from sameProductList
totalNoOfProduct=[]

sameProductList=[
        {
            "code":{"B07N7PK9QK","B09LN2XKKQ","B07QLHFSFP"},
            "name":"MIELLE 1X (OLD) - ",
            "value":0
        },
        {
            "code":{"B0BB8BFLQF","B0DJL8CQTK"},
            "name":"MIELLE OIL 3X SET - ",
            "value":0
        },
        {
            "code":{"B0DKG5N8BN","B0CJM8546H"},
            "name":"MIELLE OIL LIGHT - ",
            "value":0
        },
        {
            "code":{"B0DJDP7KJD","B0DK246TWW"},
            "name":"MIELLE SHAMPOO & CONDITIONER SET - ",
            "value":0
        },
        {
            "code":{"B01HOD3W78","B0038TXGL0"},
            "name":"SM CH SHAMPOO - ",
            "value":0
        },
        {
            "code":{"B01HOD3W4Q","B00EUMC62O"},
            "name":"SM CH CONDITIONER - ",
            "value":0
        },
        {
            "code":{"B01HOD3ZQG","B00MXDBWI6"},
            "name":"SM JBCO MASQUE - ",
            "value":0
        },
        {
            "code":{"B00OVQO66S","B07XWWJ1X1"},
            "name":"SM MH CONDITIONER - ",
            "value":0
        },
        {
            "code":{"B07NSZHMWD","B07XLZHCN7"},
            "name":"SM MH MASQUE - ",
            "value":0
        },
        {
            "code":{"B0DFQFYHKC","B0DKTVTM5N"},
            "name":"MILLI ROSEMARY OIL - ",
            "value":0
        },

     ]

# Iterate over the rows in the sheet (assuming headers in the first row)
for row in sheet.iter_rows(min_row=2, values_only=True):
    code = row[0]  # Column A (code)
    name = row[1]  # Column B (name)
    value = row[2]  # Column C (value)

    product_matched = False
  # Check if code matches any of the sameProductList groups  
    for product in sameProductList:
        if code in product["code"]:
            if value is None:
                value = 0
            product["value"] += value #if isinstance(value, (int, float)) else 0  # Add value to the matched product
            totalProduct+=value
            product["code"].remove(code)
            if len(product["code"]) == 0:
                result_list.append({
                "code": " ",
                "name": product["name"],
                "value": product["value"]
            })
            product_matched = True
            break    
     

    if not product_matched:
        # Check if any of the values in the row are null
        if code is None and name is None and value is None:
            code = " "
            name = " "  
            value = " "
            totalNoOfProduct.append(totalProduct)
            totalProduct=0 
        elif code is not None and name is None and value is not None :
            code = " "
            name = row[0]
            value = totalProduct   
        elif code is not None and name is None and value is None:
            code = " "
            name = row[0] 
            value = " "
        elif value is None:
            value =" "
        else:
            totalProduct += value
        
        if name=="AUSTRELIA" or name=="FRANCE" or name=="GERMANY"  or name=="ITALY" or name=="UK":
            result_list.append({
                "code":" ",
                "name":f"{name} (NEW) - {row[3]} (OLD) - {row[5]} = {row[3]+row[5]}",
                "value":" "
            })
        elif name=="TOTAL PRODUCT OF PCS :":
             result_list.append({
                "code":" ",
                "name":name,
                "value":sum(totalNoOfProduct)
            })
        else:    
            # If no match, add the row data to the result_list
            result_list.append({
                "code": code,
                "name": name,
                "value": value
            })

#print(result_list)
# Create a new text file and write the result_list data into it
output_file_path = "C:/All Sales/Sales/DAILY WHATSAPP MSG.txt"  # Path for the output text file

try:
    with open(output_file_path, "w") as file:
        for item in result_list:
            # # Write each item as a formatted string
            file.write(f"{item['name']} {item['value']}\n")
        
    print(f"Data successfully written to {output_file_path}")
except Exception as e:
    print(f"An error occurred: {e}")

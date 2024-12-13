#  one store inventory like NW, KC, SP, JM  individually


import glob
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta
from pathlib import Path

folder_path = "C:/All Sales/Sales/KC"
# Load the workbook and specify the sheet name   ........Destination excel file
file_path = "C:/All Sales/Sales/Master Inventory.xlsx"  # Path to your Excel file
workbook = load_workbook(file_path)
sheet = workbook.active


# # Use glob to find all Excel files (with .xls and .xlsx extensions)
excel_files = glob.glob(f"{folder_path}/*.csv")
df=[]
for data in excel_files:
   df.append(pd.read_csv(data))

# Step 2: Initialize result_list as an empty list
result_list = []
# Iterate over the rows in the sheet (assuming headers in the first row)
for row in sheet.iter_rows(min_row=2, values_only=True):
    code = row[0]  # Column A (code)
    name = row[1]  # Column B (name)
    value = row[2]  # Column C (value)

 # Check if any of the values in the row are null
    if code is None or name is None or value is None:
        # print(f"Row contains null values: {row}")
        continue  # Skip processing this row if you want to ignore it

    # Append data in the desired format
    result_list.append({
        "code": code,
         "name": name,
         "value": 0
    })

# Print the result list to see the data
# print(result_list)

# Step 4: Compare result_list codes with df2 codes and perform operations
for dff in df[0:]:
    for _, row2 in dff.iterrows():
        code_found = False

        # Check if code from df2 exists in result_list
        for item in result_list:
            if item["code"] == row2['(Child) ASIN']:
                # If code matches, perform addition
                # if item["code"]=="B00128WK4I":
                #     print(row2['Units ordered'])
                item["value"] += row2['Units ordered']
                # print('match found', row2['(Child) ASIN'],item["value"])
                code_found = True
                break

        # If code not found, add the new code and value to result_list
        # if not code_found:
        #     result_list.append({"code": row2['(Child) ASIN'], "value": row2['Units ordered']})


# print(result_list)


def get_value_by_code(code):
    # Iterate over the result_list and check for the code
    for item in result_list:
        if item["code"] == code:
            return item["value"]
    return None  # Return None if code is not found

for code in result_list:
    # Define the value to search for
    search_value = code["code"]
    # Initialize a variable to store the cell address
    cell_address = None


    # Iterate through the rows and columns to find the value
    for row in sheet.iter_rows():  # Iterate over all rows
        for cell in row:  # Iterate over each cell in the row
            if cell.value == search_value:  # Match the target value
                cell_address = cell.coordinate  # Get the cell address
                break  # Stop the inner loop if found
        if cell_address:
            break  # Stop the outer loop if found

    # Output the result
    if cell_address:
        # print(f"The value '{search_value}' is located at: {cell_address}")
        # Extract row number from the original address
        row_number = int(cell_address[1:])  # Extract everything after the first character (row)
        new_column = "C"  # Specify the new column

        # Construct the new cell address
        new_cell_address = f"{new_column}{row_number}"

        # Copy value from original cell to new cell
        sheet[new_cell_address] = get_value_by_code(search_value)
        workbook.save(file_path)
    else:
        print(f"The value '{search_value}' was not found in the sheet.")


#change file name like master inventory.xlsx to SP master inventory 2044-06-15.xlsx
if os.path.exists(file_path):
    # Rename the file
    today = datetime.today()
    today_str = today.strftime("%Y-%m-%d")

    #get folder name
    folder_name = Path(folder_path).name

    #rename file name
    new_file_path = f"C:/Sales/{folder_name} Master Inventory {today_str}.xlsx"
    os.rename(file_path,new_file_path)
    # print(f"File renamed from '{file_path}")
else:
    print(f"The file '{file_path}' does not exist.")

# ..................................MIELLE OLD AND NEW CLASSIFICATION........................

# Define variables to store MIELLE OLD and NEW values
regions = {"AUS": {"old": 0, "new": 0}, 
           "FRA": {"old": 0, "new": 0}, 
           "GER": {"old": 0, "new": 0}, 
           "IT": {"old": 0, "new": 0}, 
           "UK": {"old": 0, "new": 0}}

# MIELLE OLD and NEW codes
MIELLE_OLD_CODE = {"B07N7PK9QK", "B09LN2XKKQ", "B07QLHFSFP"}
MIELLE_NEW_CODE = {"B0DHVLFR2V"}

# Map file names to region keys
region_mapping = {
    "KC AUS": "AUS",
    "KC FRA": "FRA",
    "KC GER": "GER",
    "KC SPA": "GER",
    "KC SWE": "GER",
    "KC BEL": "GER",
    "KC NL": "GER",
    "KC POL": "GER",
    "KC IT": "IT",
    "KC UK": "UK",
}

# Process each file and update region counts
for file, dff in zip(excel_files, df):
    file_name = os.path.basename(os.path.splitext(file)[0])
    region_key = region_mapping.get(file_name)
    
    if region_key:
        for _, row in dff.iterrows():
            if row["(Child) ASIN"] in MIELLE_OLD_CODE:
                regions[region_key]["old"] += row["Units ordered"]
            elif row["(Child) ASIN"] in MIELLE_NEW_CODE:
                regions[region_key]["new"] += row["Units ordered"]

        
for reg in regions:
    print(f"{reg} - OLD: {regions[reg]['old']}, NEW: {regions[reg]['new']}")               
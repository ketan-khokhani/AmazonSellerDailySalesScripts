
#  calculation of one store one country and enter data into the daily sales report.xlsx


import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta

# Define multiple folder paths
folder_paths = [
    "C:/All Sales/Sales/NW",
    "C:/All Sales/Sales/KC",
    "C:/All Sales/Sales/SP",
    "C:/All Sales/Sales/JM"
]

# Define the base file path (assuming this file remains the same for all folders)
file_path = "C:/All Sales/Sales/Daily Sales Report.xlsx"

# Mapping for country-specific sections based on folder names
country_mapping_list = {
    "JM": {
        "AUS": "J M LIMITED (AUSTRALIA)",
        "UK": "J M LIMITED (UK)",
        "GER": "J M LIMITED (GERMANY)",
        "SWE": "J M LIMITED (GERMANY)",
        "BEL": "J M LIMITED (GERMANY)",
        "NL": "J M LIMITED (GERMANY)",
        "POL": "J M LIMITED (GERMANY)",
        "SPA": "J M LIMITED (GERMANY)",
        "FRA": "J M LIMITED (FRANCE)",
        "IT": "J M LIMITED (ITALY)"
    },
    "NW": {
        "AUS": "NORTH WEST (AUSTRALIA)",
        "UK": "NORTH WEST (UK)",
        "GER": "NORTH WEST (GERMANY)",
        "SWE": "NORTH WEST (GERMANY)",
        "BEL": "NORTH WEST (GERMANY)",
        "NL": "NORTH WEST (GERMANY)",
        "POL": "NORTH WEST (GERMANY)",
        "SPA": "NORTH WEST (GERMANY)",
        "FRA": "NORTH WEST (FRANCE)",
        "IT": "NORTH WEST (ITALY)"
    },
    "SP": {
        "AUS": "SPETRA (AUSTRALIA)",
        "UK": "SPETRA (UK)",
        "GER": "SPETRA (GERMANY)",
        "SWE": "SPETRA (GERMANY)",
        "BEL": "SPETRA (GERMANY)",
        "NL": "SPETRA (GERMANY)",
        "POL": "SPETRA (GERMANY)",
        "SPA": "SPETRA (GERMANY)",
        "FRA": "SPETRA (FRANCE)",
        "IT": "SPETRA (ITALY)"
    },
    "KC": {
        "AUS": "KC STORE (AUSTRALIA)",
        "UK": "KC STORE (UK)",
        "GER": "KC STORE (GERMANY)",
        "SWE": "KC STORE (GERMANY)",
        "BEL": "KC STORE (GERMANY)",
        "NL": "KC STORE (GERMANY)",
        "POL": "KC STORE (GERMANY)",
        "SPA": "KC STORE (GERMANY)",
        "FRA": "KC STORE (FRANCE)",
        "IT": "KC STORE (ITALY)"
    }
}

# Mapping for sheet names based on folder names
sheet_name_mapping = {
    "KC": "KC Product Sales",
    "NW": "NW Product Sales",
    "JM": "JM Product Sales",
    "SP": "SP Product Sales"
}

# Get yesterday's date in the required format (e.g., "11 DEC 2024")
yesterday = datetime.today() - timedelta(days=1)
yesterday_str = yesterday.strftime("%d %b %Y").upper()  # Format as "11 DEC 2024"

# Get the day of the week (e.g., "Monday")
day_of_week = yesterday.strftime("%A").upper()  # Full weekday name (e.g., "MONDAY")

# Load the workbook and iterate over each sheet to add the yesterday's date in the first row, second column
if os.path.exists(file_path):
    try:
        # Load the Excel workbook
        workbook = load_workbook(file_path)
        
        # Iterate over all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Add yesterday's date in the first row and second column
            sheet.cell(row=1, column=2).value = yesterday_str

            # Add day of the week in the second row and second column
            sheet.cell(row=2, column=2).value = day_of_week
        
        # Save the changes to the Excel file
        workbook.save(file_path)
        print(f"Yesterday's date ({yesterday_str}) and the day of the week ({day_of_week}) added to each sheet.")
    
    except Exception as e:
        print(f"Error processing the Excel file: {e}")
else:
    print(f"The file '{file_path}' does not exist.")



# Process each folder
for folder_path in folder_paths:
    print(f"Processing folder: {folder_path}")
    folder_name = Path(folder_path).name  # Extract the folder name
    country_mapping = country_mapping_list.get(folder_name)  # Get the country mapping for this folder
    matched_sheet_name = sheet_name_mapping.get(folder_name)  # Get the corresponding sheet name

    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        continue

    if not matched_sheet_name:
        print(f"Error: No matching sheet name found for folder: {folder_name}")
        continue

    # Load Excel sheet
    try:
        data = pd.read_excel(file_path, sheet_name=matched_sheet_name, engine='openpyxl')
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        continue

   
    # Read CSV files in the folder
    data_by_country = {}
    files = os.listdir(folder_path)
    for file_nm in files:
        file_pt = os.path.join(folder_path, file_nm)

        if file_nm.endswith(".csv"):
            try:
                country_name = file_nm.split()[1].split(".")[0] if len(file_nm.split()) > 1 else "Unknown"
                csv_data = pd.read_csv(file_pt)
                data_by_country[country_name] = csv_data
            except Exception as e:
                print(f"Error loading CSV file {file_nm}: {e}")

    # Update Excel with data from CSVs
    for country_code, country_data in data_by_country.items():
        section = country_mapping.get(country_code)
        if not section:
            print(f"Warning: No section mapping found for country code: {country_code}")
            continue

        # Find section in the Excel sheet
        data.iloc[:, 1] = data.iloc[:, 1].str.strip()  # Clean up whitespace
        section_indices = data[data.iloc[:, 1] == section].index

        if len(section_indices) == 0:
            print(f"Warning: Section '{section}' not found in the sheet '{matched_sheet_name}'.")
            continue

        start_row = section_indices[0] + 1
        end_row = len(data)
        for idx in range(start_row, len(data)):
            if pd.isna(data.iloc[idx, 0]):  # Look for the first blank row
                end_row = idx
                break

        # Update cells in the Excel file
        section_data = data.iloc[start_row:end_row].reset_index(drop=True)
        for _, row in country_data.iterrows():
            target_code = row['(Child) ASIN']
            unit_ordered = row['Units ordered']

            found_row = section_data[section_data.iloc[:, 0].str.strip() == target_code]
            if not found_row.empty:
                relative_row_index = found_row.index[0]
                absolute_row_index = start_row + relative_row_index + 2

                try:
                    workbook = load_workbook(file_path)
                    sheet = workbook[matched_sheet_name]
                    cell_value = sheet.cell(row=absolute_row_index, column=6).value
                    # Check if the cell has a value
                    if cell_value is not None and cell_value != "":
                        sheet.cell(row=absolute_row_index, column=6).value += unit_ordered
                    else:
                       sheet.cell(row=absolute_row_index, column=6).value = unit_ordered
                    
                    workbook.save(file_path)
                except Exception as e:
                    print(f"Error updating Excel file: {e}")
                    continue

                #print(f"Updated code '{target_code}' in section '{section}' at row {absolute_row_index} and value {unit_ordered}.")
            else:
                print(f"Code '{target_code}' not found in section '{section}'.")

#change file name like master inventory.xlsx to SP master inventory 2044-06-15.xlsx
if os.path.exists(file_path):
    # Rename the file
    today = datetime.today()
    yesterday = datetime.today() - timedelta(days=1)
    yesterday_str = yesterday.strftime("%Y-%m-%d")

    #get file name
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    print(file_name)
    #rename file name
    new_file_path = f"C:/All Sales/Sales/{file_name} {yesterday_str}.xlsx"
    os.rename(file_path,new_file_path)
    # print(f"File renamed from '{file_path}")
else:
    print(f"The file '{file_path}' does not exist.")
